#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
แดชบอร์ดคัดกรองสุขภาพช่องปากนักเรียน — Standalone app
อ่านข้อมูลจากไฟล์ .xlsm/.xlsx (ชีต "Records") ในโฟลเดอร์เดียวกับตัวโปรแกรม
สร้างหน้าแดชบอร์ด (HTML) แล้วเปิดในเบราว์เซอร์ พร้อมตรวจสอบเวอร์ชันใหม่จาก GitHub

Build เป็น .exe ด้วย PyInstaller (ดู .github/workflows/release.yml)
"""
import os, sys, json, math, glob, tempfile, webbrowser, urllib.request, traceback

# ---- ค่าที่ workflow จะแทนที่ตอน build (อย่าเปลี่ยนรูปแบบสตริง) ----
__version__ = "0.0.0"
GITHUB_REPO = "OWNER/REPO"      # เช่น "nchotenin/oral-health-dashboard"
# ------------------------------------------------------------------

# ================= การถอดรหัสค่า (บาก์อินในแอป) =================
DEC = {
 'sex':{'1':'ชาย','2':'หญิง'},
 'nutrition_status':{'1':'ปกติ','2':'ท้วม/อ้วน','3':'ผอม/ค่อนข้างผอม'},
 'caregiver_main':{'1':'พ่อแม่','2':'ปู่ย่าตายาย','3':'พี่เลี้ยง','4':'อื่นๆ'},
 'caregiver_caries6m':{'0':'ไม่ใช่','1':'ใช่'},
 'siblings_caries':{'0':'ไม่ใช่','1':'ใช่'},
 'brush_morning':{'0':'ไม่เคย','1':'บางวัน','2':'ทุกวัน'},
 'brush_after_lunch':{'0':'ไม่เคย','1':'บางวัน','2':'ทุกวัน'},
 'brush_before_bed':{'0':'ไม่เคย','1':'บางวัน','2':'ทุกวัน'},
 'eat_drink_after_bed':{'0':'ไม่','1':'ใช่','2':'อื่นๆ'},
 'brush_duration':{'1':'< 2 นาที','2':'≥ 2 นาที','3':'ไม่ทราบ'},
 'toothpaste_use':{'0':'ไม่ใช้','1':'ใช้'},
 'toothpaste_f':{'1':'มีฟลูออไรด์','2':'ไม่มีฟลูออไรด์'},
 'cleaning_aid_use':{'1':'ใช้','2':'ไม่ใช้'},
 'sweet_drink_freq':{'0':'ไม่ดื่ม','1':'ดื่ม 1-2 ครั้ง/วัน','2':'ดื่ม 3-4 ครั้ง/วัน','3':'ดื่ม 5 ครั้ง/วัน ขึ้นไป'},
 'snack_freq':{'0':'ไม่กิน','1':'กิน 1-2 ครั้ง/วัน','2':'กิน 3-4 ครั้ง/วัน','3':'กิน 5 ครั้ง/วัน ขึ้นไป'},
 'candy_freq':{'0':'ไม่กิน','1':'บางวัน','2':'ทุกวัน'},
 'add_sugar_freq':{'1':'ไม่เติม','2':'บางครั้ง','3':'ทุกครั้ง'},
 'milk_habit':{'0':'ไม่ดื่ม','1':'ดื่มนมรสจืดเป็นประจำ','2':'ดื่มนมรสหวาน/เปรี้ยว/ปรุงแต่งรส','3':'ดื่มทั้งนมจืดและนมปรุงแต่งรส','4':'ดื่มทั้งนมจืดและนมปรุงแต่งรส'},
 'smoking_status':{'0':'ไม่เคยสูบ','1':'เคยลองสูบ ปัจจุบันเลิกแล้ว','2':'ยังสูบจนถึงปัจจุบัน'},
 'toothache_ever':{'0':'ไม่เคย','1':'เคย'},
 'bad_breath':{'0':'ไม่เคย','1':'นานๆครั้ง','2':'เป็นประจำ','9':'ไม่แน่ใจ/ไม่ทราบ'},
 'bleeding_brush':{'0':'ไม่เคย','1':'นานๆครั้ง','2':'เป็นประจำ','9':'ไม่แน่ใจ/ไม่ทราบ'},
 'food_impaction':{'0':'ไม่เคย','1':'นานๆครั้ง','2':'เป็นประจำ','9':'ไม่แน่ใจ/ไม่ทราบ'},
 'dentine_sensitivity':{'0':'ไม่เคย','1':'นานๆครั้ง','2':'เป็นประจำ','9':'ไม่แน่ใจ/ไม่ทราบ'},
 'gingival_abscess':{'0':'ไม่เคย','1':'นานๆครั้ง','2':'เป็นประจำ','9':'ไม่แน่ใจ/ไม่ทราบ'},
 'dental_check_12m':{'0':'ไม่เคย','1':'เคย ครูตรวจให้','2':'เคย หมอฟันตรวจให้','3':'เคย ทั้งครูและหมอฟัน'},
 'dental_treat_12m':{'0':'ไม่เคย','1':'เคย'},
 'oral_appliance':{'0':'ไม่มี','1':'มี'},
 'got_oh_education':{'0':'ไม่ได้รับ','1':'ได้รับ'},
 'gi':{'0':'ปกติ','1':'เหงือกอักเสบ','9':'ไม่ประเมิน'},
 'ohplaque':{'0':'สะอาด','1':'มีคราบจุลินทรีย์','9':'ไม่ประเมิน'},
 'is_fluorosis':{'0':'ไม่มี','1':'มี (Fluorosis)'},
}
_YESNO01 = {'0':'ไม่','1':'ใช่'}
for _f in ['place_school_dentist','place_health_center','place_public_hosp','place_private',
           'edu_online','edu_radio','edu_loudspeaker','edu_poster','edu_tv','edu_parents',
           'edu_teacher','edu_vhv','edu_health_staff','edu_friends','edu_books']:
    DEC[_f] = _YESNO01

# ค่าที่ควร "เฝ้าระวัง" (แสดงแท่งสีแดง + ⚠️) — ใช้เมื่อไม่มีชีต Codebook
ALERT_DEFAULT = {
 'caregiver_caries6m':['1'],'siblings_caries':['1'],'eat_drink_after_bed':['1'],'brush_duration':['1'],
 'brush_morning':['0'],'brush_after_lunch':['0'],'brush_before_bed':['0'],'toothpaste_f':['2'],
 'cleaning_aid_use':['2'],'sweet_drink_freq':['2','3'],'snack_freq':['2','3'],'candy_freq':['2'],
 'add_sugar_freq':['3'],'milk_habit':['2','4'],'smoking_status':['2'],'toothache_ever':['1'],
 'bad_breath':['2'],'bleeding_brush':['2'],'food_impaction':['2'],'dentine_sensitivity':['2'],
 'gingival_abscess':['2'],'dental_check_12m':['0'],'is_fluorosis':['1'],'gi':['1'],'ohplaque':['1'],
 'r_defect':['1'],'r_caries':['3'],'r_vplaque':['3'],'r_flu':['1'],'r_ohplaque':['3'],
 'r_brushflu':['1'],'r_app':['2'],'r_sweet':['2'],'r_siblings':['1'],
}
EXCLUDE = {'สภาวะฟัน','สภาวะเหงือก','คราบจุรินทรีย์','คราบจุลินทรีย์','แฟ็ม Dental','แฟ็มDental'}
CLASS_ORDER = ['อ.2','อ.3','ป.1','ป.2','ป.3','ป.4','ป.5','ป.6','6','ม.1','ม.2','ม.3']
NUM_FIELDS = {'age_years','age_months','weight_kg','height_cm','money_per_day','money_snack_per_day',
              'candy_times_per_day','candy_pieces','absent_days_toothache','absent_days_treatment',
              'teeth','d','m','f','goodteeth','riskteeth','sealant','fracture','riskscore',
              'r_defect','r_caries','r_vplaque','r_flu','r_ohplaque','r_brushflu','r_app','r_sweet','r_siblings'}
DROP = {'round_id','year','school_id','cid','name','interviewer_name','interview_date','birth_date',
        'measure_source','caregiver_other','toothpaste_brand_other','toothpaste_brand_local',
        'cleaning_aid_text','oral_appliance_text','edu_other_text','smoke_start_age',
        'defect','absent_for_dentist','othercon'}
CAR = {'K','P','1','2'}


def resource_path(rel):
    """หา path ของไฟล์ที่ฝังมากับ exe (PyInstaller) หรืออยู่ข้าง ๆ สคริปต์"""
    base = getattr(sys, '_MEIPASS', os.path.dirname(os.path.abspath(__file__)))
    return os.path.join(base, rel)


def app_dir():
    """โฟลเดอร์ที่ตัว exe/สคริปต์อยู่ (ใช้ค้นหาไฟล์ข้อมูล)"""
    if getattr(sys, 'frozen', False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))


def find_data_workbook(folder):
    """หาไฟล์ .xlsm/.xlsx ที่มีชีต 'Records' ในโฟลเดอร์"""
    import openpyxl
    cands = sorted(glob.glob(os.path.join(folder, '*.xlsm')) +
                   glob.glob(os.path.join(folder, '*.xlsx')))
    cands = [c for c in cands if not os.path.basename(c).startswith('~$')]
    for c in cands:
        try:
            wb = openpyxl.load_workbook(c, read_only=True, data_only=True)
            if 'Records' in wb.sheetnames:
                wb.close()
                return c
            wb.close()
        except Exception:
            continue
    return None


def _norm_school(v):
    if v is None: return None
    return 'โรงเรียนบ้านซำตะเคียน' if 'ซำตะเคียน' in str(v) else str(v)

def _norm_risk(v):
    if v is None: return None
    s = str(v).strip()
    return 'เสี่ยงต่ำ' if s in ('ต่ำ','เสี่ยงต่ำ') else s


def load_and_transform(path):
    """อ่านเวิร์กบุ๊ก แล้วสร้าง payload สำหรับหน้าแดชบอร์ด (โครงเดียวกับที่ dashboard ต้องใช้)"""
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb['Records']
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 4:
        raise ValueError("ชีต 'Records' มีข้อมูลไม่ครบ (ต้องมีหัวตาราง 3 แถว + ข้อมูล)")

    # แถว 0 = หมวดหมู่ (เติมค่าลงมา), แถว 1 = ชื่อไทย, แถว 2 = code, แถว 3+ = ข้อมูล
    cat_row, thai_row, code_row = rows[0], rows[1], rows[2]
    codes, cat_of, thai_of, order = [], {}, {}, []
    cur_cat = None
    for i, code in enumerate(code_row):
        cat = cat_row[i] if i < len(cat_row) else None
        if cat not in (None, ''):
            cur_cat = str(cat).strip()
        c = str(code).strip() if code not in (None, '') else None
        codes.append(c)
        if c:
            cat_of[c] = cur_cat
            thai_of[c] = thai_row[i] if i < len(thai_row) else None
            if c not in order:
                order.append(c)

    data = []
    for r in rows[3:]:
        if all(v is None for v in r):
            continue
        rec = {}
        for c, v in zip(codes, r):
            if c:
                rec[c] = v
        data.append(rec)

    # alert: ใช้ Codebook ถ้ามี ไม่งั้นใช้ค่าเริ่มต้น
    alert_map = dict(ALERT_DEFAULT)
    if 'Codebook' in wb.sheetnames:
        cb = wb['Codebook']
        for row in list(cb.iter_rows(values_only=True))[1:]:
            if not row or row[0] in (None, ''):
                continue
            name = str(row[0]).strip()
            al = row[4] if len(row) > 4 else None
            if al not in (None, ''):
                alert_map[name] = [x.strip() for x in str(al).split(',')]

    included_cats, seen = [], set()
    for c in order:
        cat = cat_of.get(c)
        if cat and cat not in EXCLUDE and cat not in seen:
            included_cats.append(cat); seen.add(cat)
    included_fields = [c for c in order if cat_of.get(c) not in EXCLUDE and c not in DROP]
    TEETH = [c for c in order if cat_of.get(c) == 'สภาวะฟัน' and c != 'fluorosis']

    def tv(v):
        return '' if v in (None, '') else str(v).strip()

    recs = []
    for r in data:
        o = {f: r.get(f) for f in included_fields}
        o['school_name'] = _norm_school(r.get('school_name'))
        o['class_level'] = str(r.get('class_level')) if r.get('class_level') is not None else None
        o['riskgroup'] = _norm_risk(r.get('riskgroup'))
        o['round_id'] = str(r.get('round_id')) if r.get('round_id') is not None else None
        o['year'] = str(r.get('year')) if r.get('year') is not None else None
        o['name'] = r.get('name')
        o['cid'] = r.get('cid')
        dd = r.get('d') or 0; mm = r.get('m') or 0; ff = r.get('f') or 0
        val = lambda x: (x if isinstance(x, (int, float)) else 0)
        o['dmft'] = val(dd) + val(mm) + val(ff)
        o['caries_free'] = 1 if (isinstance(dd, (int, float)) and dd == 0) else 0
        # ต่อซี่ฟัน
        ncar, lst = 0, []
        for f in TEETH:
            v = tv(r.get(f)); o[f] = v
            if v in CAR:
                ncar += 1; lst.append(f[1:] + ':' + v)
        o['n_caries'] = ncar
        o['caries_list'] = ', '.join(lst)
        recs.append(o)

    def clean(v):
        if isinstance(v, float) and math.isnan(v):
            return None
        return v
    for o in recs:
        for k in list(o):
            o[k] = clean(o[k])

    meta_fields = []
    for f in included_fields:
        al = alert_map.get(f)
        meta_fields.append({
            'code': f, 'thai': thai_of.get(f), 'cat': cat_of.get(f),
            'kind': 'num' if f in NUM_FIELDS else 'enum',
            'decode': DEC.get(f),
            'alert': (None if al in (None, '') else (al if isinstance(al, list) else [al])),
        })

    def skey(x):
        return (len(str(x)), str(x))
    rounds = sorted({o['round_id'] for o in recs if o['round_id']}, key=skey)
    years = sorted({o['year'] for o in recs if o['year']}, key=skey)

    wb.close()
    return {'records': recs, 'fields': meta_fields, 'categories': included_cats,
            'class_order': CLASS_ORDER, 'rounds': rounds, 'years': years, 'teeth': TEETH}


def check_update():
    """ตรวจสอบเวอร์ชันล่าสุดจาก GitHub Releases (คืนค่า dict ถ้ามีของใหม่)"""
    if not GITHUB_REPO or GITHUB_REPO == "OWNER/REPO":
        return None
    try:
        url = f"https://api.github.com/repos/{GITHUB_REPO}/releases/latest"
        req = urllib.request.Request(url, headers={'User-Agent': 'oral-health-dashboard'})
        with urllib.request.urlopen(req, timeout=5) as resp:
            info = json.load(resp)
        latest = (info.get('tag_name') or '').lstrip('vV')
        html = info.get('html_url') or f"https://github.com/{GITHUB_REPO}/releases/latest"
        def vt(s):
            out = []
            for p in str(s).split('.'):
                n = ''.join(ch for ch in p if ch.isdigit())
                out.append(int(n) if n else 0)
            return tuple(out)
        if latest and vt(latest) > vt(__version__):
            return {'latest': latest, 'url': html}
    except Exception:
        pass
    return None


def build_html(payload):
    tpl = open(resource_path('template.html'), encoding='utf-8').read()
    logo = open(resource_path('logo_datauri.txt'), encoding='utf-8').read().strip()
    data = json.dumps(payload, ensure_ascii=False).replace('</', '<\\/')
    return tpl.replace('__DATA__', data).replace('__LOGO__', logo)


def main():
    folder = app_dir()
    print("=" * 60)
    print(" แดชบอร์ดคัดกรองสุขภาพช่องปากนักเรียน  v%s" % __version__)
    print("=" * 60)
    print(" โฟลเดอร์: %s" % folder)

    wb_path = find_data_workbook(folder)
    if not wb_path:
        print("\n [ไม่พบไฟล์ข้อมูล]")
        print(" กรุณาวางไฟล์ .xlsm หรือ .xlsx ที่มีชีตชื่อ 'Records'")
        print(" ไว้ในโฟลเดอร์เดียวกับโปรแกรมนี้ แล้วเปิดใหม่อีกครั้ง")
        input("\n กด Enter เพื่อปิด...")
        return
    print(" ไฟล์ข้อมูล: %s" % os.path.basename(wb_path))

    try:
        payload = load_and_transform(wb_path)
    except Exception as e:
        print("\n [อ่านข้อมูลไม่สำเร็จ] %s" % e)
        traceback.print_exc()
        input("\n กด Enter เพื่อปิด...")
        return
    print(" อ่านข้อมูลสำเร็จ: %d ระเบียน" % len(payload['records']))

    upd = check_update()
    payload['app'] = {'version': __version__, 'update': upd}
    if upd:
        print(" ** มีเวอร์ชันใหม่ %s — %s" % (upd['latest'], upd['url']))

    html = build_html(payload)
    out_path = os.path.join(folder, 'oral_health_dashboard.html')
    try:
        with open(out_path, 'w', encoding='utf-8') as f:
            f.write(html)
    except Exception:
        # เขียนในโฟลเดอร์โปรแกรมไม่ได้ → ใช้ temp
        fd, out_path = tempfile.mkstemp(suffix='.html', prefix='oral_dashboard_')
        with os.fdopen(fd, 'w', encoding='utf-8') as f:
            f.write(html)

    print(" เปิดแดชบอร์ดในเบราว์เซอร์...")
    webbrowser.open('file:///' + out_path.replace('\\', '/'))
    print(" เสร็จสิ้น (หน้าต่างนี้ปิดได้)")


if __name__ == '__main__':
    try:
        main()
    except Exception:
        traceback.print_exc()
        input("\n เกิดข้อผิดพลาด — กด Enter เพื่อปิด...")
